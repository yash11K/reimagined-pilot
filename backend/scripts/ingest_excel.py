"""Ingest an Excel export (e.g. Prod_Decagon_KB_All.xlsx) into KB Manager.

Workflow per row:
    1. MetadataEnricher (Haiku) derives title, filename, brand, category,
       tags, visibility from the raw content.
    2. Source record created  (type=manual, url=decagon://<id>)
    3. IngestionJob created   (status → completed)
    4. KBFile created         (status → approved)
    5. Markdown + metadata sidecar uploaded to S3

The result is identical in schema to AEM-ingested files.

Logs are written live to logs/ingest_excel_<timestamp>.log for tracking.
A JSON summary report is saved to logs/ingest_excel_<timestamp>_summary.json.

Usage:
    python -m scripts.ingest_excel                        # full run
    python -m scripts.ingest_excel --dry-run              # preview only
    python -m scripts.ingest_excel --skip-s3              # DB only
    python -m scripts.ingest_excel --concurrency 20       # faster
    python -m scripts.ingest_excel --file other.xlsx      # custom file
"""

from __future__ import annotations

import argparse
import asyncio
import json
import logging
import sys
import time
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

# Ensure project root is on sys.path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from kb_manager.agents.metadata_enricher import EnrichedMetadata, MetadataEnricher
from kb_manager.config import get_settings
from kb_manager import database as db_module
from kb_manager.queries import files as file_queries
from kb_manager.queries import jobs as job_queries
from kb_manager.queries import sources as source_queries
from kb_manager.services.s3_uploader import S3Uploader

logger = logging.getLogger(__name__)

# ── Logging setup ───────────────────────────────────────────────────────

LOG_DIR = Path(__file__).resolve().parent.parent / "logs"
RUN_TS = datetime.now().strftime("%Y%m%d_%H%M%S")


def setup_logging() -> Path:
    """Configure dual logging: console + live log file. Returns log file path."""
    LOG_DIR.mkdir(exist_ok=True)
    log_file = LOG_DIR / f"ingest_excel_{RUN_TS}.log"

    fmt = logging.Formatter(
        "%(asctime)s  %(levelname)-7s  %(message)s", datefmt="%H:%M:%S"
    )

    # Console handler
    console = logging.StreamHandler()
    console.setFormatter(fmt)
    console.setLevel(logging.INFO)

    # File handler — flush every line for live tailing
    file_handler = logging.FileHandler(str(log_file), encoding="utf-8")
    file_handler.setFormatter(
        logging.Formatter("%(asctime)s  %(levelname)-7s  %(message)s")
    )
    file_handler.setLevel(logging.INFO)

    root = logging.getLogger()
    root.setLevel(logging.INFO)
    root.addHandler(console)
    root.addHandler(file_handler)

    logger.info("📝 Log file: %s", log_file)
    return log_file

# ── Excel column indices (0-based) ──────────────────────────────────────
COL_ID = 0            # Decagon ID
COL_DISPLAY_NAME = 1  # Display Name
COL_CONTENT = 5       # Content (Q&A markdown)
COL_TAGS = 6          # Tags — brand / category hints
COL_URL = 7           # URL (rarely populated)
COL_CREATED = 8       # Created At
COL_UPDATED = 9       # Updated At

DEFAULT_URL = "https://www.avis.com/en/home"


def parse_iso(val) -> datetime | None:
    """Best-effort ISO timestamp parse."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.replace(tzinfo=timezone.utc) if val.tzinfo is None else val
    try:
        dt = datetime.fromisoformat(str(val))
        return dt.replace(tzinfo=timezone.utc) if dt.tzinfo is None else dt
    except Exception:
        return None


def load_excel(path: Path) -> list[dict]:
    """Load Excel rows into a list of dicts with named keys."""
    wb = openpyxl.load_workbook(str(path), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        content = row[COL_CONTENT]
        if not content:
            continue
        rows.append({
            "decagon_id": row[COL_ID],
            "display_name": row[COL_DISPLAY_NAME],
            "content": str(content),
            "tags_raw": row[COL_TAGS],
            "url": row[COL_URL],
            "created_at": parse_iso(row[COL_CREATED]),
            "updated_at": parse_iso(row[COL_UPDATED]),
        })
    wb.close()
    return rows


# ── Enrich a single row via Haiku ───────────────────────────────────────

async def enrich_row(
    enricher: MetadataEnricher,
    row: dict,
    idx: int,
) -> EnrichedMetadata:
    """Call the MetadataEnricher for one row, with retry."""
    try:
        return await enricher.run(
            row["content"],
            tags_hint=row["tags_raw"],
            display_name=row["display_name"],
        )
    except Exception as exc:
        logger.warning("⚠️ Row %d enrichment failed (%s) — using fallback", idx, exc)
        return EnrichedMetadata(
            title=str(row["display_name"] or "Untitled FAQ"),
            filename=f"faq-{row['decagon_id'] or idx}",
            brand="unknown",
            category="faq",
        )


# ── Persist a single row to DB + S3 ────────────────────────────────────

async def persist_row(
    row: dict,
    meta: EnrichedMetadata,
    idx: int,
    s3: S3Uploader | None,
    kb_target: str,
    region: str,
) -> str:
    """Create Source → Job → KBFile → S3. Returns status string."""
    source_url = row["url"] or DEFAULT_URL
    decagon_url = f"decagon://{row['decagon_id']}" if row["decagon_id"] else f"decagon://row-{idx}"

    async with db_module.async_session_factory() as db:
        # 1. Source
        source = await source_queries.create_source(
            db,
            type="manual",
            url=decagon_url,
            region=region,
            brand=meta.brand,
            kb_target=kb_target,
            metadata_={
                "decagon_id": row["decagon_id"],
                "display_name": row["display_name"],
                "tags_raw": row["tags_raw"],
                "original_url": row["url"],
            },
        )

        # 2. Job
        job = await job_queries.create_job(
            db,
            source_id=source.id,
            status="processing",
        )

        # 3. KBFile
        kb_file = await file_queries.create_file(
            db,
            job_id=job.id,
            title=meta.title,
            md_content=row["content"],
            source_url=source_url,
            region=region,
            brand=meta.brand,
            kb_target=kb_target,
            category=meta.category,
            visibility=meta.visibility,
            tags=meta.tags if meta.tags else None,
            status="approved",
            modify_date=row["updated_at"] or row["created_at"],
        )

        # 4. Link source ↔ file
        await file_queries.link_source_to_file(db, source.id, kb_file.id)

        # 5. S3 upload
        s3_key = None
        if s3 is not None:
            s3_key = s3.upload(kb_file)
            if s3_key:
                await file_queries.update_file(db, kb_file.id, s3_key=s3_key)

        # 6. Complete job & mark source ingested
        await job_queries.update_job_status(db, job.id, "completed")
        await source_queries.mark_ingested(db, source.id)

        await db.commit()

    return "✅" if s3_key else "⚠️ (no S3)"


# ── Main orchestrator ────────────────────────────────────────────────────

async def run(args: argparse.Namespace) -> None:
    log_file = setup_logging()
    run_start = time.perf_counter()

    excel_path = Path(args.file)
    if not excel_path.exists():
        logger.error("File not found: %s", excel_path)
        sys.exit(1)

    rows = load_excel(excel_path)
    if args.limit > 0:
        rows = rows[: args.limit]
    logger.info("📊 Loaded %d rows from '%s'", len(rows), excel_path.name)

    # ── Tracking state ───────────────────────────────────────────────
    brand_counter: Counter = Counter()
    category_counter: Counter = Counter()
    visibility_counter: Counter = Counter()
    enrichment_fallbacks = 0
    persist_success = 0
    persist_failed = 0
    s3_uploaded = 0
    s3_skipped = 0
    failed_rows: list[dict] = []

    # ── Phase 1: Enrich metadata via Haiku (concurrent) ─────────────
    enriched: list[tuple[int, dict, EnrichedMetadata]] = []
    cache_file = LOG_DIR / f"ingest_excel_{RUN_TS}_enrichment_cache.json"

    if args.resume:
        # Load from a previous enrichment cache
        resume_path = Path(args.resume)
        if not resume_path.exists():
            logger.error("Resume cache not found: %s", resume_path)
            sys.exit(1)
        logger.info("♻️  Resuming from enrichment cache: %s", resume_path)
        cached = json.loads(resume_path.read_text(encoding="utf-8"))
        # Rebuild enriched list from cache + rows
        row_map = {r["decagon_id"]: r for r in rows}
        for entry in cached:
            idx = entry["row"]
            meta = EnrichedMetadata(
                title=entry["title"],
                filename=entry["filename"],
                brand=entry["brand"],
                category=entry["category"],
                visibility=entry.get("visibility", "public"),
                tags=entry.get("tags", []),
            )
            # Match row by decagon_id
            row = row_map.get(entry.get("decagon_id"))
            if row is None:
                # Fallback: match by position
                row_idx = idx - 2  # rows are 0-indexed, idx starts at 2
                row = rows[row_idx] if 0 <= row_idx < len(rows) else None
            if row is None:
                logger.warning("⚠️ Could not match cache entry row %d — skipping", idx)
                continue
            if meta.brand == "unknown" and meta.category == "general":
                enrichment_fallbacks += 1
            enriched.append((idx, row, meta))
        enriched.sort(key=lambda x: x[0])
        enrich_elapsed = 0.0
        logger.info("♻️  Loaded %d enriched rows from cache", len(enriched))
    else:
        # Run enrichment via Haiku
        enricher = MetadataEnricher()
        semaphore = asyncio.Semaphore(args.concurrency)
        enrich_start = time.perf_counter()

        async def _enrich_with_limit(idx: int, row: dict):
            nonlocal enrichment_fallbacks
            async with semaphore:
                meta = await enrich_row(enricher, row, idx)
                if meta.brand == "unknown" and meta.category == "general":
                    enrichment_fallbacks += 1
                enriched.append((idx, row, meta))

        logger.info("🏷️  Phase 1: Enriching metadata (%d concurrent)...", args.concurrency)
        tasks = [_enrich_with_limit(i, row) for i, row in enumerate(rows, start=2)]
        await asyncio.gather(*tasks)
        enriched.sort(key=lambda x: x[0])  # restore row order

        enrich_elapsed = time.perf_counter() - enrich_start
        logger.info(
            "🏷️  Enrichment complete — %d rows in %.1fs (%.1f rows/sec, %d fallbacks)",
            len(enriched), enrich_elapsed,
            len(enriched) / enrich_elapsed if enrich_elapsed > 0 else 0,
            enrichment_fallbacks,
        )

        # Save enrichment cache
        cache_data = [
            {
                "row": idx,
                "decagon_id": row["decagon_id"],
                "title": meta.title,
                "filename": meta.filename,
                "brand": meta.brand,
                "category": meta.category,
                "visibility": meta.visibility,
                "tags": meta.tags,
            }
            for idx, row, meta in enriched
        ]
        cache_file.write_text(json.dumps(cache_data, indent=2, default=str), encoding="utf-8")
        logger.info("💾 Enrichment cache saved: %s", cache_file)

    # Collect enrichment stats
    for _, _, meta in enriched:
        brand_counter[meta.brand] += 1
        category_counter[meta.category] += 1
        visibility_counter[meta.visibility] += 1

    # ── Dry-run: just print ──────────────────────────────────────────
    if args.dry_run:
        for idx, row, meta in enriched:
            logger.info(
                "[DRY-RUN] Row %d | %s | brand=%s | cat=%s | vis=%s | tags=%s | file=%s",
                idx, meta.title[:60], meta.brand, meta.category,
                meta.visibility, meta.tags, meta.filename,
            )
        _log_summary(args, len(enriched), 0, 0, 0, 0, enrichment_fallbacks,
                      brand_counter, category_counter, visibility_counter,
                      enrich_elapsed, 0, failed_rows, log_file, dry_run=True)
        return

    # ── Phase 2: Persist to DB + S3 (sequential for DB safety) ──────
    db_module.init_engine()

    s3: S3Uploader | None = None
    if not args.skip_s3:
        try:
            s3 = S3Uploader()
        except Exception as e:
            logger.warning("⚠️ S3Uploader init failed (%s) — skipping uploads", e)

    persist_start = time.perf_counter()
    logger.info("💾 Phase 2: Persisting to DB + S3...")

    for idx, row, meta in enriched:
        try:
            status = await persist_row(row, meta, idx, s3, args.kb_target, args.region)
            if "✅" in status:
                s3_uploaded += 1
            else:
                s3_skipped += 1
            logger.info("%s Row %d → %s | %s.md", status, idx, meta.title[:50], meta.filename)
            persist_success += 1
        except Exception as exc:
            logger.exception("❌ Row %d failed to persist", idx)
            persist_failed += 1
            failed_rows.append({
                "row": idx,
                "decagon_id": row["decagon_id"],
                "title": meta.title,
                "error": str(exc),
            })

    persist_elapsed = time.perf_counter() - persist_start
    await db_module.dispose_engine()

    total_elapsed = time.perf_counter() - run_start
    _log_summary(args, len(enriched), persist_success, persist_failed,
                  s3_uploaded, s3_skipped, enrichment_fallbacks,
                  brand_counter, category_counter, visibility_counter,
                  enrich_elapsed, persist_elapsed, failed_rows, log_file)

    logger.info("Total runtime: %.1fs", total_elapsed)


def _log_summary(
    args,
    total_rows: int,
    success: int,
    failed: int,
    s3_uploaded: int,
    s3_skipped: int,
    enrichment_fallbacks: int,
    brand_counter: Counter,
    category_counter: Counter,
    visibility_counter: Counter,
    enrich_elapsed: float,
    persist_elapsed: float,
    failed_rows: list[dict],
    log_file: Path,
    dry_run: bool = False,
) -> None:
    """Log a summary to console/file and write a JSON report."""
    logger.info("━" * 60)
    logger.info("INGESTION SUMMARY")
    logger.info("━" * 60)
    logger.info("Source file:          %s", args.file)
    logger.info("Mode:                 %s", "DRY-RUN" if dry_run else "LIVE")
    logger.info("Total rows:           %d", total_rows)
    logger.info("Enrichment fallbacks: %d", enrichment_fallbacks)
    logger.info("Enrichment time:      %.1fs", enrich_elapsed)

    if not dry_run:
        logger.info("DB persisted:         %d ✅  %d ❌", success, failed)
        logger.info("S3 uploaded:          %d ✅  %d skipped", s3_uploaded, s3_skipped)
        logger.info("Persist time:         %.1fs", persist_elapsed)

    logger.info("─" * 40)
    logger.info("Brand breakdown:")
    for brand, count in brand_counter.most_common():
        logger.info("  %-20s %d", brand, count)

    logger.info("Category breakdown:")
    for cat, count in category_counter.most_common():
        logger.info("  %-20s %d", cat, count)

    logger.info("Visibility breakdown:")
    for vis, count in visibility_counter.most_common():
        logger.info("  %-20s %d", vis, count)

    if failed_rows:
        logger.info("─" * 40)
        logger.info("Failed rows:")
        for fr in failed_rows:
            logger.info("  Row %d (decagon=%s): %s — %s",
                        fr["row"], fr["decagon_id"], fr["title"], fr["error"][:100])

    logger.info("━" * 60)

    # Write JSON summary report
    summary = {
        "run_timestamp": RUN_TS,
        "source_file": args.file,
        "mode": "dry_run" if dry_run else "live",
        "total_rows": total_rows,
        "enrichment": {
            "fallbacks": enrichment_fallbacks,
            "elapsed_seconds": round(enrich_elapsed, 1),
        },
        "persist": {
            "success": success,
            "failed": failed,
            "s3_uploaded": s3_uploaded,
            "s3_skipped": s3_skipped,
            "elapsed_seconds": round(persist_elapsed, 1),
        },
        "brands": dict(brand_counter.most_common()),
        "categories": dict(category_counter.most_common()),
        "visibility": dict(visibility_counter.most_common()),
        "failed_rows": failed_rows,
    }

    summary_file = LOG_DIR / f"ingest_excel_{RUN_TS}_summary.json"
    summary_file.write_text(json.dumps(summary, indent=2, default=str), encoding="utf-8")
    logger.info("📋 Summary report: %s", summary_file)


def main():
    parser = argparse.ArgumentParser(
        description="Ingest Excel export into KB Manager with LLM-enriched metadata"
    )
    parser.add_argument("--file", default="Prod_Decagon_KB_All.xlsx",
                        help="Path to Excel file")
    parser.add_argument("--kb-target", default="public",
                        help="KB target (default: public)")
    parser.add_argument("--region", default="US",
                        help="Region (default: US)")
    parser.add_argument("--concurrency", type=int, default=10,
                        help="Max concurrent Haiku calls (default: 10)")
    parser.add_argument("--limit", type=int, default=0,
                        help="Process only first N rows (0 = all)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Enrich metadata but skip DB/S3 writes")
    parser.add_argument("--skip-s3", action="store_true",
                        help="Write to DB only, skip S3 upload")
    parser.add_argument("--resume", type=str, default=None,
                        help="Path to enrichment cache JSON to skip Haiku and resume persist")
    args = parser.parse_args()
    asyncio.run(run(args))


if __name__ == "__main__":
    main()
